from flask import Flask , render_template, request
from openpyxl import load_workbook
app= Flask(__name__)

@app.route("/")
def index():
    images = []
    excel= load_workbook("gallery.xlsx")
    ws= excel["Лист1"]
    for row in ws:
        url = row[0].value
        title = row[2].value
        lst = [title,  url] 
        images.append(lst)

    return render_template("index.html",images=images )


@app.route("/add")
def add():

    return render_template("add.html")

@app.route("/reciever" , methods=["POST"])
def reciever():
    url = request.form.get("url") 
    title = request.form.get("title")
    excel= load_workbook("gallery.xlsx")
    sheet = excel["Лист1"]
    sheet.append([ url , description, title]) 
    excel.save("gallery.xlsx")
   
    return render_template("add.html") 


@app.route("/details/<number>")
def details(number):
    excel = load_workbook("gallery.xlsx")
    page = excel["Лист1"]    
    lst = page[str(number)]
    return render_template("details.html", lst=lst)